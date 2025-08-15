import sys
from pathlib import Path

sys.path.append(str(Path(__file__).resolve().parents[1]))

from services import exports
from openpyxl import load_workbook


def sample_records():
    return [
        {
            "name": "Alice",
            "email": "alice@example.com",
            "journal": "J",
            "topic": "T",
            "verified": True,
            "duplicate": False,
            "source_url": "http://example.com",
            "timestamp": "2024-01-01T00:00:00",
        }
    ]


def test_to_csv():
    csv_text = exports.to_csv(sample_records())
    lines = csv_text.strip().splitlines()
    assert lines[0].startswith("name,email")
    assert "alice@example.com" in lines[1]


def test_to_excel_file(tmp_path):
    path = tmp_path / "out.xlsx"
    exports.to_excel_file(sample_records(), path)
    wb = load_workbook(path)
    ws = wb.active
    assert ws.title == "data"
    assert ws.max_row == 2
    headers = [cell.value for cell in ws[1]]
    assert headers[0] == "name"
